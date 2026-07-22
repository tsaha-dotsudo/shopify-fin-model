import pandas as pd

SHOP_NAME = "SHOP_NAME"  # set your store name here
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

raw = "raw/"
pnl = pd.read_csv(raw+"pnl_monthly.csv")
cust = pd.read_csv(raw+"customers_monthly.csv")
fun = pd.read_csv(raw+"funnel_monthly.csv")
prod = pd.read_csv(raw+"products.csv")

wb = Workbook()

ARIAL = "Arial"
BLUE = Font(name=ARIAL, size=10, color="0000FF")
BLACK = Font(name=ARIAL, size=10)
GREEN = Font(name=ARIAL, size=10, color="008000")
HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="1A1A1A")
YELLOW = PatternFill("solid", fgColor="FFFF00")
TITLE = Font(name=ARIAL, size=13, bold=True)
BOLD = Font(name=ARIAL, size=10, bold=True)
THIN = Border(bottom=Side(style="thin", color="CCCCCC"))

INR = u'\u20B9#,##0;(\u20B9#,##0);-'
INR2 = u'\u20B9#,##0.00;(\u20B9#,##0.00);-'
PCT = '0.0%;(0.0%);-'
NUM = '#,##0;(#,##0);-'

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR; c.fill = HDRFILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

# ---------- README ----------
ws = wb.active; ws.title = "README"
ws["A1"] = f"{SHOP_NAME} Financial Model"; ws["A1"].font = TITLE
lines = [
    "",
    f"Source: Shopify ShopifyQL ({SHOP_NAME}), pulled 21 Jul 2026. Period: Oct 2025 to Jul 2026. Jul 2026 is PARTIAL (through 21 Jul).",
    "",
    "Sheets:",
    "  Data_PnL / Data_Customers / Data_Funnel / Data_Products = raw Shopify report data. Do not edit.",
    "  Assumptions = your cost inputs. Fill the yellow cells.",
    "  Model = monthly financial model, all formulas. Recalculates when Assumptions change.",
    "  Insights = number-backed findings as of this pull.",
    "",
    "Legend: blue text = hardcoded input you can change. Yellow fill = fill these in with your real costs.",
    "Black = formula. Green = pulls from another sheet.",
    "",
    "Refresh: re-run the same 4 ShopifyQL queries (listed in row 16-19), paste new rows into the Data sheets, extend Model columns.",
    "",
    "Q1: FROM sales SHOW orders, gross_sales, discounts, returns, net_sales, shipping_charges, taxes, total_sales TIMESERIES month SINCE 2025-10-01 UNTIL today",
    "Q2: FROM sales SHOW orders, gross_sales, discounts, returns, net_sales GROUP BY product_title SINCE 2025-10-01 UNTIL today ORDER BY net_sales DESC LIMIT 30",
    "Q3: FROM sales SHOW customers, returning_customers, returning_customer_rate, average_order_value TIMESERIES month SINCE 2025-10-01 UNTIL today",
    "Q4: FROM sessions SHOW sessions, sessions_with_cart_additions, sessions_that_reached_checkout, sessions_that_completed_checkout, conversion_rate TIMESERIES month SINCE 2025-10-01 UNTIL today",
]
for i, t in enumerate(lines, 2):
    ws.cell(row=i, column=1, value=t).font = BLACK
ws.column_dimensions["A"].width = 140

# ---------- Data sheets ----------
def dump(df, name, money_cols, pct_cols=()):
    ws = wb.create_sheet(name)
    header_row(ws, 1, list(df.columns), [14]+[13]*(len(df.columns)-1) if name!="Data_Products" else [52]+[12]*(len(df.columns)-1))
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BLACK; cell.border = THIN
            col = df.columns[c-1]
            if col in money_cols: cell.number_format = INR2
            elif col in pct_cols: cell.number_format = PCT
            elif isinstance(v, (int, float)): cell.number_format = NUM
    return ws

dump(pnl, "Data_PnL", {"gross_sales","discounts","returns","net_sales","shipping_charges","taxes","total_sales"})
dump(cust, "Data_Customers", {"average_order_value"}, {"returning_customer_rate"})
dump(fun, "Data_Funnel", set(), {"conversion_rate"})
dump(prod, "Data_Products", {"gross_sales","discounts","returns","net_sales"})

# ---------- Assumptions ----------
ws = wb.create_sheet("Assumptions")
ws["A1"] = "Cost Assumptions (fill yellow cells with your real numbers)"; ws["A1"].font = TITLE
rows = [
    ("COGS % of net sales (materials: TPU/nylon/silicone, clasps, filament)", 0.30, PCT, "Placeholder example. Replace with your blended unit cost / price."),
    ("Courier cost per order (Blue Dart / Delivery Express avg)", 90, INR, "Placeholder example. Replace with your actual avg per shipment."),
    ("Packaging cost per order (hang-sell pack, shrink wrap, label)", 25, INR, "Placeholder example."),
    ("Payment gateway % of total sales (Razorpay)", 0.02, PCT, "Standard Razorpay domestic rate. Adjust if on a different plan."),
    ("Monthly fixed costs (Shopify Basic, apps, tools)", 3500, INR, "Placeholder example. Shopify Basic + misc."),
]
header_row(ws, 3, ["Assumption", "Value", "Note"], [62, 14, 70])
for i, (label, val, fmt, note) in enumerate(rows, 4):
    ws.cell(row=i, column=1, value=label).font = BLACK
    c = ws.cell(row=i, column=2, value=val); c.font = BLUE; c.fill = YELLOW; c.number_format = fmt
    ws.cell(row=i, column=3, value=note).font = BLACK
ws["A10"] = "All values above are examples to show format, not real costs. Model outputs marked 'est.' depend on them."
ws["A10"].font = Font(name=ARIAL, size=9, italic=True)

# ---------- Model ----------
ws = wb.create_sheet("Model")
ws["A1"] = "Monthly Model  (Jul 2026 partial, through 21 Jul)"; ws["A1"].font = TITLE
months = list(pnl["month"])
n = len(months)
header_row(ws, 3, ["Metric"] + months, [40] + [12]*n)
for i, m in enumerate(months):
    ws.cell(row=3, column=2+i).value = m

def moneyrow(r, label, formula_fn, fmt=INR, font=GREEN, bold=False):
    c = ws.cell(row=r, column=1, value=label)
    c.font = BOLD if bold else BLACK
    for i in range(n):
        col = get_column_letter(2+i)
        dcol = get_column_letter(2+i)  # data sheets: col A=month, then B.. ; data row = i+2
        cell = ws.cell(row=r, column=2+i, value=formula_fn(i, i+2))
        cell.font = Font(name=ARIAL, size=10, bold=bold, color=font.color) if not isinstance(font, str) else BLACK
        cell.number_format = fmt
        cell.border = THIN

# Data_PnL columns: A month B orders C gross D discounts E returns F net G shipping H taxes I total
moneyrow(4,  "Orders",              lambda i,dr: f"=Data_PnL!B{dr}", NUM)
moneyrow(5,  "Gross sales",         lambda i,dr: f"=Data_PnL!C{dr}")
moneyrow(6,  "Discounts",           lambda i,dr: f"=Data_PnL!D{dr}")
moneyrow(7,  "Returns",             lambda i,dr: f"=Data_PnL!E{dr}")
moneyrow(8,  "Net sales",           lambda i,dr: f"=Data_PnL!F{dr}", INR, GREEN, True)
moneyrow(9,  "Shipping charged",    lambda i,dr: f"=Data_PnL!G{dr}")
moneyrow(10, "Total sales",         lambda i,dr: f"=Data_PnL!I{dr}")

ws.cell(row=12, column=1, value="Growth & quality").font = BOLD
def fml(r, label, fn, fmt):
    ws.cell(row=r, column=1, value=label).font = BLACK
    for i in range(n):
        col = get_column_letter(2+i)
        cell = ws.cell(row=r, column=2+i, value=fn(i, i+2, col))
        cell.font = BLACK; cell.number_format = fmt; cell.border = THIN

fml(13, "Net sales MoM growth", lambda i,dr,col: "" if i==0 else f"=IFERROR({col}8/{get_column_letter(1+i)}8-1,\"\")".replace("{col}8", f"{get_column_letter(2+i)}8"), PCT)
fml(14, "Discount rate (% of gross)", lambda i,dr,col: f"=IFERROR(-{col}6/{col}5,0)", PCT)
fml(15, "Return rate (% of gross)",   lambda i,dr,col: f"=IFERROR(-{col}7/{col}5,0)", PCT)
fml(16, "AOV (net sales / order)",    lambda i,dr,col: f"=IFERROR({col}8/{col}4,0)", INR)
fml(17, "Shipping charged / order",   lambda i,dr,col: f"=IFERROR({col}9/{col}4,0)", INR)

ws.cell(row=19, column=1, value="Traffic & conversion").font = BOLD
# Data_Funnel: A month B sessions C cart D reached E completed F conv
fml(20, "Sessions",                lambda i,dr,col: f"=Data_Funnel!B{dr}", NUM)
fml(21, "Conversion rate",         lambda i,dr,col: f"=Data_Funnel!F{dr}", PCT)
fml(22, "Cart adds / session",     lambda i,dr,col: f"=IFERROR(Data_Funnel!C{dr}/Data_Funnel!B{dr},0)", PCT)
fml(23, "Checkout completion (completed/reached)", lambda i,dr,col: f"=IFERROR(Data_Funnel!E{dr}/Data_Funnel!D{dr},0)", PCT)
fml(24, "Revenue per session (total sales)", lambda i,dr,col: f"=IFERROR({col}10/{col}20,0)", INR)

ws.cell(row=26, column=1, value="Customers").font = BOLD
# Data_Customers: A month B customers C returning D rate E aov
fml(27, "Customers",               lambda i,dr,col: f"=Data_Customers!B{dr}", NUM)
fml(28, "Returning customer rate", lambda i,dr,col: f"=Data_Customers!D{dr}", PCT)

ws.cell(row=30, column=1, value="Contribution (est., driven by Assumptions)").font = BOLD
A = "Assumptions"
fml(31, "COGS (est.)",            lambda i,dr,col: f"=-{col}8*{A}!$B$4", INR)
fml(32, "Courier cost (est.)",    lambda i,dr,col: f"=-{col}4*{A}!$B$5", INR)
fml(33, "Packaging (est.)",       lambda i,dr,col: f"=-{col}4*{A}!$B$6", INR)
fml(34, "Gateway fees (est.)",    lambda i,dr,col: f"=-{col}10*{A}!$B$7", INR)
fml(35, "Fixed costs",            lambda i,dr,col: f"=-{A}!$B$8", INR)
def contrib(r, label, fn, fmt, bold=True):
    ws.cell(row=r, column=1, value=label).font = BOLD
    for i in range(n):
        col = get_column_letter(2+i)
        cell = ws.cell(row=r, column=2+i, value=fn(col))
        cell.font = BOLD; cell.number_format = fmt; cell.border = THIN
contrib(36, "Contribution (est.)",   lambda col: f"={col}8+{col}9+SUM({col}31:{col}35)", INR)
contrib(37, "Contribution margin %", lambda col: f"=IFERROR({col}36/{col}8,0)", PCT)

ws.freeze_panes = "B4"

# ---------- Insights ----------
ws = wb.create_sheet("Insights")
ws["A1"] = "Insights — data as of 21 Jul 2026"; ws["A1"].font = TITLE
insights = [
("1. GROWTH", [
 "Net sales grew from ₹62.6k (Nov, first full month) to ₹460.8k (Jun) — 7.4x in 7 months, avg ~33% MoM.",
 "Jul is tracking a dip: ₹267.0k net in 21 days = ~₹395k run-rate vs ₹460.8k in Jun (~-14%).",
]),
("2. THE JULY DIP IS TRAFFIC, NOT CONVERSION", [
 "Jul conversion is your best ever: 3.16% vs 2.51% in Jun. Revenue per session ₹40.2 vs ₹34.2 in Jun (+18%).",
 "But sessions run-rate is ~9,979/mo vs 13,752 in Jun (-27%). The store converts better than ever; fewer people are reaching it.",
 "Action: this is a top-of-funnel problem. Push content/ads/SEO, not site changes.",
]),
("3. SHIPPING INCOME COLLAPSED IN MAY-JUL", [
 "Shipping charged per order: Apr ₹101 → May ₹48 → Jun ₹26 → Jul ₹38.",
 "If courier cost is ~₹90/order, Jun alone absorbed ~₹24k of unrecovered shipping (376 orders x ~₹64 gap).",
 "Action: verify this was intentional (free shipping threshold?). If margin is tight, raise the free-shipping cart minimum.",
]),
("4. REPEAT BUSINESS IS COMPOUNDING", [
 "Returning customer rate: 8.3% (Nov) → 21% (May-Jun) → 27.8% (Jul).",
 "Nearly 1 in 3.6 July customers is a repeat buyer. Bands are consumable/collectible — email flows and colour drops are working or would work harder.",
]),
("5. PRODUCT CONCENTRATION", [
 "Top 5 SKUs = ₹564k net = ~31% of all-time net sales. Helio Bicep line alone ≈ ₹513k+ across colours.",
 "Adapter V3 (Pack of 3): 174 orders (most-ordered SKU) but only ₹26.4k net (₹152/order). It's an entry product — bundle it with bands to lift AOV.",
 "Whoop 4.0 Band Black: ₹116.8k net but also the most discounted SKU (₹2,953). Check if discounts here are necessary.",
]),
("6. DISCOUNTS AND RETURNS ARE UNDER CONTROL", [
 "Discount rate peaked at 3.7% of gross (Mar-Apr, Summer Saving tiers) and is down to 2.2% in Jul.",
 "Return rate has never exceeded 3.1% (Dec) and sits at 0.7% in Jul. Strict 24h defective-only policy is holding.",
]),
("7. AOV IS FLAT — THE GROWTH LEVER LEFT UNTOUCHED", [
 "AOV has hovered ₹1,014-1,241 for 9 months with no trend. All growth has come from order volume.",
 "At 3.16% conversion, +₹100 AOV on Jun volume = +₹37.6k/mo. Levers: adapter+band bundles, tiered free shipping, 2-pack colour offers.",
]),
]
r = 3
for head, pts in insights:
    ws.cell(row=r, column=1, value=head).font = BOLD; r += 1
    for p in pts:
        ws.cell(row=r, column=1, value="   - " + p).font = BLACK; r += 1
    r += 1
ws.column_dimensions["A"].width = 150

wb.save(f"{SHOP_NAME.replace(' ', '_')}_Financial_Model.xlsx")
print("saved")
